from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from database.db import get_db_connection
from utils.pdf_utils import extract_text_chunks_from_pdf
from utils.embeddings import embed_and_store_with_page
from utils.embeddings import model, collection
from flask import request




app = Flask(__name__)
app.secret_key = "healthapp_secret_key"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# ---------------- USER CLASS ----------------
class User(UserMixin):
    def __init__(self, id, username, password):
        self.id = id
        self.username = username
        self.password = password

@login_manager.user_loader
def load_user(user_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id, username, password FROM Users WHERE id=?",
        (user_id,)
    )
    row = cursor.fetchone()
    conn.close()
    if row:
        return User(row[0], row[1], row[2])
    return None

# -------------------- Excel Q&A Function to generate answer (USED BY BOTH ASK + EXCEL) ------------------------------
# ================= STEP 2 =================
# Function to generate answer (USED BY BOTH ASK + EXCEL)

import subprocess
def generate_answer_for_question(question):
    # 1️⃣ Encode question
    query_embedding = model.encode(question).tolist()

    # 2️⃣ Retrieve relevant chunks
    results = collection.query(
        query_embeddings=[query_embedding],
        n_results=3
    )

    context = ""
    sources = []

    # ✅ SAFETY CHECK (prevents ZeroDivisionError also)
    if not results["documents"] or not results["documents"][0]:
        return "No relevant information found.", 0.0, []

    for doc_text, metadata in zip(results["documents"][0], results["metadatas"][0]):
        context += doc_text + "\n"
        sources.append({
            "doc": metadata.get("filename", "Unknown document"),
            "page": metadata.get("page", "N/A")
        })

    prompt = f"""
Answer the question using ONLY the provided context.
If the answer is not found, say "Not specified in the document."

Context:
{context}

Question:
{question}
"""

    try:
        answer = subprocess.check_output(
            ["ollama", "run", "tinyllama:latest", prompt],
            text=True
        )
    except Exception:
        answer = "Model failed to generate answer."
# ----NEW-----
    distances = results.get("distances", [[]])[0]
    if distances:
        similarities = [max(0, 1 - d) for d in distances]
        confidence = round(
            (0.6 * similarities[0] + 0.4 * (sum(similarities) / len(similarities))),
            2
            )
        confidence = round(confidence * 100, 2)
    else:
        confidence = 0.0

# ----Old logic-----
    # distances = results.get("distances", [[]])[0]
    # if distances:
    #     similarities = [1 - d for d in distances]   # cosine similarity
    #     confidence = round(sum(similarities) / len(similarities), 2)
    # else:
    #     confidence = 0.0
    #confidence = round(1 - (sum(distances) / len(distances)), 2) if distances else 0.0

    # ✅ RETURN LIST, NOT STRING
    return answer.strip(), confidence, sources

# --------------------Old generate answers route -------------------    

# def generate_answer_for_question(question):
#     # 1️⃣ Encode question
#     query_embedding = model.encode(question).tolist()

#     # 2️⃣ Retrieve relevant chunks
#     results = collection.query(
#         query_embeddings=[query_embedding],
#         n_results=3
#     )

#     context = ""
#     sources = []

#     for doc_text, metadata in zip(results["documents"][0], results["metadatas"][0]):
#         context += doc_text + "\n"
#         sources.append(f"{metadata['filename']} (Page {metadata['page']})")

#     prompt = f"""
# Answer the question using ONLY the provided context.
# If the answer is not found, say "Not specified in the document."

# Context:
# {context}

# Question:
# {question}
# """

#     try:
#         answer = subprocess.check_output(
#             ["ollama", "run", "tinyllama:latest", prompt],
#             text=True
#         )
#     except Exception:
#         answer = "Model failed to generate answer."

#     distances = results.get("distances", [])

#     if distances and len(distances[0]) > 0:
#         avg_distance = sum(distances[0]) / len(distances[0])
#         confidence = round(1 - avg_distance, 2)
#     else:
#         confidence = 0.0

    # distances = results.get("distances", [[0.5]])[0]  ---------------------keep this commented----
    # confidence = round(1 - (sum(distances) / len(distances)), 2) ---------------------keep this commented----

   # return answer.strip(), confidence, ", ".join(sources)


# ---------------- ROUTES ----------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        email = request.form["email"]
        password = generate_password_hash(request.form["password"])

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Users (username, email, password) VALUES (?, ?, ?)",
            (username, email, password)
        )
        conn.commit()
        conn.close()

        return redirect(url_for("login"))

    return render_template("register.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, username, password FROM Users WHERE username=?",
            (username,)
        )
        row = cursor.fetchone()
        conn.close()

        if row and check_password_hash(row[2], password):
            user = User(row[0], row[1], row[2])
            login_user(user)
            return redirect(url_for("dashboard"))
        else:
            return "Invalid credentials"

    return render_template("login.html")

# --------- Dashboard Route ----------------
# @app.route("/dashboard")
# @login_required
# def dashboard():
#     return render_template("dashboard.html")
@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch uploaded documents for this user
    cursor.execute("""
        SELECT id, filename, upload_time
        FROM Documents
        WHERE user_id = ?
        ORDER BY upload_time DESC
    """, (current_user.id,))

    docs = cursor.fetchall()
    conn.close()

    documents = []
    for d in docs:
        documents.append({
            "id": d[0],
            "name": d[1],
            "date": d[2]
        })

    return render_template(
        "dashboard.html",
        documents=documents
    )




@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))

# # ---------- Upload route ---------
# import os
# from flask import flash

# UPLOAD_FOLDER = "uploads"
# app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# @app.route("/upload", methods=["GET", "POST"])
# @login_required
# def upload_pdf():
#     if request.method == "POST":
#         file = request.files["pdf"]

#         if file.filename.endswith(".pdf"):
#             filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
#             file.save(filepath)

#             conn = get_db_connection()
#             cursor = conn.cursor()
#             cursor.execute(
#                 "INSERT INTO Documents (user_id, filename) VALUES (?, ?)",
#                 (current_user.id, file.filename)
#             )
#             conn.commit()
#             conn.close()

#             return "PDF uploaded successfully! <a href='/dashboard'>Go back</a>"

#         return "Only PDF files allowed"

#     return render_template("upload.html")


# ----------------PDF Upload Route -----------------------
import os
from flask import request, render_template
from database.db import get_db_connection
from flask_login import login_required, current_user

UPLOAD_FOLDER = "uploads/pdf"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)  # create folder if it doesn't exist
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

@app.route("/upload_pdf", methods=["GET", "POST"])
@login_required
def upload_pdf():
    if request.method == "POST":
        file = request.files.get("pdf")
        if file and file.filename.endswith(".pdf"):
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], file.filename)
            file.save(filepath)
           
              # 2️⃣ Extract text from PDF
            chunks_with_page = extract_text_chunks_from_pdf(filepath)

            # 3️⃣ Store embeddings with page info in ChromaDB
            embed_and_store_with_page(filename=file.filename, chunks_with_page=chunks_with_page)
            #embed_and_store(file.filename, chunks_with_page)

            # Save metadata in database
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO Documents (user_id, filename) VALUES (?, ?)",
                (current_user.id, file.filename)
            )
            conn.commit()
            conn.close()

            # ✅ SUCCESS MESSAGE
            flash("✅ PDF uploaded successfully!")
            return redirect(url_for("upload_pdf"))
        # ❌ ERROR MESSAGE
        flash("❌ Only PDF files are allowed!")


           # return "PDF uploaded successfully! <a href='/dashboard'>Go back</a>"

       # return "Only PDF files are allowed!"

    return render_template("upload_pdf.html")


# ---------------- Add Excel Upload Route NEWWWW -----------------
import openpyxl
import os

@app.route("/upload_excel", methods=["GET", "POST"])
@login_required
def upload_excel():
    recent_results = []   # 👈 ONLY THIS UPLOAD

    if request.method == "POST":
        file = request.files.get("excel")

        if file and file.filename.endswith(".xlsx"):
            filepath = os.path.join("uploads/excel", file.filename)
            file.save(filepath)

            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            conn = get_db_connection()
            cursor = conn.cursor()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                question = row[0]
                if not question:
                    continue

                # 1️⃣ Insert Question
                cursor.execute(
                    "INSERT INTO Questions (user_id, question_text) OUTPUT INSERTED.id VALUES (?, ?)",
                    (current_user.id, question)
                )
                question_id = cursor.fetchone()[0]

                # 2️⃣ Generate Answer
                answer, confidence, sources = generate_answer_for_question(question)

                # ✅ Convert sources to string
                if isinstance(sources, list):
                    sources_str = ", ".join(
                        f"{s['doc']} (Page {s['page']})" for s in sources
                        )
                else:
                    sources_str = sources

                # 3️⃣ Insert Answer
                cursor.execute(
                    """
                    INSERT INTO Answers
                    (question_id, user_id, answer_text, confidence_score, source_document)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (question_id, current_user.id, answer, confidence, sources_str)
                )

                # 🔹 STORE ONLY CURRENT UPLOAD RESULTS
                recent_results.append({
                    "question": question,
                    "answer": answer,
                    "confidence": confidence,
                    "source": sources_str
                })

            conn.commit()
            conn.close()

            flash("✅ Excel uploaded and answers generated successfully!")

    return render_template(
        "upload_excel.html",
        recent_results=recent_results
    )


# -------------------------------------------------Old--------------------------------------
# @app.route("/upload_excel", methods=["GET", "POST"])
# @login_required
# def upload_excel():
#     if request.method == "POST":
#         file = request.files.get("excel")

#         if file and file.filename.endswith(".xlsx"):
#             os.makedirs("uploads/excel", exist_ok=True)
#             filepath = os.path.join("uploads/excel", file.filename)
#             file.save(filepath)

#             # Read Excel
#             workbook = openpyxl.load_workbook(filepath)
#             sheet = workbook.active

#             conn = get_db_connection()
#             cursor = conn.cursor()

#             inserted = 0
#             answered = 0

#             for row in sheet.iter_rows(min_row=2, values_only=True):
#                 question = row[0]

#                 if question:
#                     # 1️⃣ Save question
#                     cursor.execute(
#                         "INSERT INTO Questions (user_id, question_text) VALUES (?, ?)",
#                         (current_user.id, str(question))
#                     )

#                     # 2️⃣ Get inserted question ID
#                     cursor.execute("SELECT @@IDENTITY")
#                     question_id = cursor.fetchone()[0]

#                     # 3️⃣ Generate answer using RAG
#                     answer, confidence, sources = generate_answer_for_question(question)

#                     # Convert sources to string
#                     source_docs = ", ".join(
#                         [f"{s['doc']} (Page {s['page']})" for s in sources]
#                     )

#                     # 4️⃣ Save answer
#                     cursor.execute(
#                         """
#                         INSERT INTO Answers
#                         (question_id, user_id, answer_text, confidence_score, source_document)
#                         VALUES (?, ?, ?, ?, ?)
#                         """,
#                         (question_id, current_user.id, answer, confidence, source_docs)
#                     )

#                     inserted += 1
#                     answered += 1

#             conn.commit()
#             conn.close()

#             flash(f"✅ Excel uploaded: {inserted} questions | {answered} answers generated")
#             return redirect(url_for("upload_excel"))

#         flash("❌ Only Excel (.xlsx) files allowed")
#         return redirect(url_for("upload_excel"))

#     return render_template("upload_excel.html")


#   ------------- Q&A route -------------------------------
# @app.route("/ask", methods=["GET", "POST"])
# @login_required
# def ask_question():
#     answer = None
#     sources = []  # To store doc name + page numbers

#     if request.method == "POST":
#         question = request.form.get("question")

#         # 1️⃣ Encode the question
#         query_embedding = model.encode(question).tolist()

#         # 2️⃣ Retrieve top 3 most relevant chunks from ChromaDB
#         results = collection.query(
#             query_embeddings=[query_embedding],
#             n_results=3
#         )

#         # 3️⃣ Combine context and keep track of document + page
#         context = ""
#         for doc_text, metadata in zip(results['documents'][0], results['metadatas'][0]):
#             context += doc_text + "\n"
#             sources.append({
#                 "doc": metadata["doc_id"],
#                 "page": metadata["page"]
#             })

#         # 4️⃣ Send to Ollama (local LLM) for answer
#         import subprocess
#         cmd = f"ollama generate healthmodel '{question}\nContext:{context}'"
#         response = subprocess.check_output(cmd, shell=True, text=True)
#         answer = response.strip()

#     return render_template("ask.html", answer=answer, sources=sources)


# ---------------------- History Route -------------------------------------------
# @app.route("/history")
# @login_required
# def history():
#     return render_template("history.html")

#--------------------------------------Old history route 17-01-26-------------
# @app.route("/history")
# @login_required
# def history():
#     conn = get_db_connection()
#     cursor = conn.cursor()

#     cursor.execute("""
#         SELECT 
#             U.username,
#             Q.question_text,
#             A.answer_text,
#             A.confidence_score,
#             A.source_document,
#             A.created_at
#         FROM Answers A
#         JOIN Questions Q ON A.question_id = Q.id
#         JOIN Users U ON A.user_id = U.id
#         ORDER BY A.created_at DESC
#     """)

#     history_data = cursor.fetchall()
#     conn.close()

#     return render_template("history.html", history_data=history_data)
@app.route("/history")
@login_required
def history():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            a.id AS answer_id,
            q.question_text,
            a.answer_text,
            a.confidence_score,
            a.source_document,
            a.is_correct,
            a.created_at
        FROM Answers a
        JOIN Questions q ON a.question_id = q.id
        WHERE a.user_id = ?
        ORDER BY a.created_at DESC
    """, (current_user.id,))

    rows = cursor.fetchall()
    conn.close()

    history_data = []

    for row in rows:
        history_data.append({
            "answer_id": row[0],
            "question": row[1],
            "answer": row[2],
            "confidence": row[3],
            "source": row[4],
            "is_correct": row[5],
            "date": row[6]
        })

    return render_template("history.html", history=history_data)






# ----- App Route ------------------------------------------------------------- 
# @app.route("/ask", methods=["GET", "POST"])
# @login_required
# def ask_question_page():
#     answer = None
#     sources = []
#     confidence = None

#     if request.method == "POST":
#         question = request.form.get("question")

#         # ✅ ONE LINE DOES EVERYTHING
#         answer, confidence, sources = generate_answer_for_question(question)

     

#     return render_template(
#         "ask.html",
#         answer=answer,
#         sources=sources,
#         confidence=confidence
#     )

# ----------------------------------New App Route Ask Question ---------------------
@app.route("/ask", methods=["GET", "POST"])
@login_required
def ask_question_page():
    question = None      # 30-1-2026
    answer = None
    sources = []
    confidence = None

    if request.method == "POST":
        question = request.form.get("question")

        # Generate answer
        answer, confidence, sources = generate_answer_for_question(question)

        # Convert sources to string for SQL
        source_text = ", ".join(
            [f"{s['doc']} (Page {s['page']})" for s in sources]
        ) if sources else None

        conn = get_db_connection()
        cursor = conn.cursor()

        # Insert question
        cursor.execute(
            """
            INSERT INTO Questions (user_id, question_text)
            OUTPUT INSERTED.id
            VALUES (?, ?)
            """,
            (current_user.id, question)
        )
        question_id = cursor.fetchone()[0]

        # Insert answer
        cursor.execute(
            """
            INSERT INTO Answers
            (question_id, user_id, answer_text, confidence_score, source_document)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                question_id,
                current_user.id,
                answer,
                confidence,
                source_text
            )
        )

        conn.commit()
        conn.close()

    return render_template(
        "ask.html",
        question=question,  #recent add 30-1-2026
        answer=answer,
        sources=sources,
        confidence=confidence
    )

# @app.route("/ask", methods=["GET", "POST"])
# @login_required
# def ask_question_page():
#     answer = None
#     sources = []

#     if request.method == "POST":
#         question = request.form.get("question")
#         answer, confidence, sources = generate_answer_for_question(question)  #----------------------------------------------
#         # 1️⃣ Encode question
#         query_embedding = model.encode(question).tolist()


#         # 2️⃣ Retrieve relevant chunks
#         results = collection.query(
#             query_embeddings=[query_embedding],
#             n_results=3
            
            
            
#         )

#         # ---- Debug retrieved chunks ----
#         print("==== Retrieved Chunks ====")
#         for doc_text, metadata in zip(results["documents"][0], results["metadatas"][0]):
#             print("DOC:", metadata["filename"], "PAGE:", metadata["page"])
#             print(doc_text)
#             print("------------------------")




#         context = ""
#         for doc_text, metadata in zip(results["documents"][0], results["metadatas"][0]):
#             context += doc_text + "\n"
#             sources.append({
#                 "doc": metadata["filename"],
#                 "page": metadata["page"]
#             })

#         # 3️⃣ Call Ollama
#         import subprocess
#         #prompt = f"Answer the question based only on the context.\n\nContext:\n{context}\n\nQuestion:\n{question}"
#         prompt = f"""
# Answer the question using ONLY the information about the class AWSTRENGTH.
# Do not include information from other classes.
# If the answer is not found, say "Not specified in the document."

# Context:
# {context}

# Question:
# {question}
# """
#         answer = subprocess.check_output(
#             ["ollama", "run", "tinyllama:latest", prompt],
#             text=True
#         )

#     return render_template("ask.html", answer=answer, sources=sources)

# ------------------------Excel Q& A File Download Route -----------------
import pandas as pd
from flask import send_file
import os

@app.route("/download_answers")
@login_required
def download_answers():
    conn = get_db_connection()

    query = """
    SELECT 
        q.question_text,
        a.answer_text,
        a.confidence_score,
        a.source_document,
        a.created_at
    FROM Answers a
    JOIN Questions q ON a.question_id = q.id
    WHERE a.user_id = ?
    ORDER BY a.created_at DESC
    """

    df = pd.read_sql(query, conn, params=[current_user.id])
    conn.close()

    if df.empty:
        flash("❌ No answers available to download.")
        return redirect(url_for("dashboard"))

    file_path = "uploads/excel/qa_results.xlsx"
    df.to_excel(file_path, index=False)

    return send_file(file_path, as_attachment=True)


#-----------------------Answer Edit route ---------------------
@app.route("/update_answer", methods=["POST"])
@login_required
def update_answer():
    answer_id = request.form.get("answer_id")
    answer_text = request.form.get("answer_text")
    is_correct = 1 if request.form.get("is_correct") == "on" else 0

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE Answers
        SET answer_text = ?, is_correct = ?
        WHERE id = ? AND user_id = ?
    """, (answer_text, is_correct, answer_id, current_user.id))

    conn.commit()
    conn.close()

    flash("✅ Answer updated successfully")
    return redirect(url_for("history"))

# --------------------- Globle History Route --------------------------

@app.route("/global-history")
@login_required
def global_history():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            a.id AS answer_id,
            u.username,
            q.question_text,
            a.answer_text,
            a.confidence_score,
            a.source_document,
            a.is_correct,
            a.created_at
        FROM Answers a
        JOIN Questions q ON a.question_id = q.id
        JOIN Users u ON a.user_id = u.id
        ORDER BY a.created_at DESC
    """)

    rows = cursor.fetchall()
    conn.close()

    history_data = []

    for row in rows:
        history_data.append({
            "answer_id": row[0],
            "username": row[1],
            "question": row[2],
            "answer": row[3],
            "confidence": row[4],
            "source": row[5],
            "is_correct": row[6],
            "date": row[7]
        })

    return render_template("global_history.html", history=history_data)

if __name__ == "__main__":
    app.run(debug=True)
